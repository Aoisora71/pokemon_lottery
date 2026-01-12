import time
import re
import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
import random
from webdriver_manager.chrome import ChromeDriverManager
import requests
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

from main import get_service, list_messages, get_message

# Enable ANSI color codes on Windows 10+
if sys.platform == 'win32':
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    except:
        pass  # If it fails, continue without colors

CAPTCHA_API_KEY = os.getenv("CAPTCHA_API_KEY")
EMAIL = None  # Email must be loaded from Excel file, not from .env
PASSWORD = os.getenv("PASSWORD")  # Password can be from .env or Excel file
LOGIN_URL = "https://www.pokemoncenter-online.com/lottery/login.html"
APPLY_URL = "https://www.pokemoncenter-online.com/lottery/apply.html"

# Logger callback - will be set by app.py
_logger = None
_stop_check = None
_max_lotteries = 1  # Default number of lotteries to process

def set_logger(logger_func):
    """Set the logging function to use instead of print()"""
    global _logger
    _logger = logger_func

def set_stop_check(stop_check_func):
    """Set the stop check function to check if bot should stop"""
    global _stop_check
    _stop_check = stop_check_func

def set_max_lotteries(count):
    """Set the maximum number of lotteries to process"""
    global _max_lotteries
    _max_lotteries = count

def check_stop():
    """Check if bot should stop. Raises StopIteration if stopped."""
    if _stop_check and not _stop_check():
        raise StopIteration("Bot stopped by user")

def log(message, level='info'):
    """Log a message using the logger callback if available, and always print to terminal with detailed formatting"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Color codes for terminal output (ANSI escape codes)
    colors = {
        'info': '\033[36m',      # Cyan
        'success': '\033[32m',   # Green
        'warning': '\033[33m',   # Yellow
        'error': '\033[31m',     # Red
    }
    reset_color = '\033[0m'
    bold = '\033[1m'
    
    # Check if terminal supports colors (try-except for safety)
    use_colors = True
    try:
        # Test if colors work
        if sys.platform == 'win32':
            # Windows might not support colors in all terminals
            use_colors = sys.stdout.isatty()
    except:
        use_colors = False
    
    color = colors.get(level, colors['info']) if use_colors else ''
    reset = reset_color if use_colors else ''
    bold_prefix = bold if use_colors else ''
    
    level_prefix = level.upper().ljust(8)
    
    # Format: [TIMESTAMP] [LEVEL] message
    if use_colors:
        terminal_message = f"[{timestamp}] [{bold_prefix}{color}{level_prefix}{reset}] {message}"
    else:
        # Fallback without colors
        terminal_message = f"[{timestamp}] [{level_prefix}] {message}"
    
    # If logger callback is set (from app.py), it will handle terminal output
    # Otherwise, print directly to terminal (when running bot.py directly)
    if _logger:
        # Web app mode: logger callback handles terminal output
        _logger(message, level)
    else:
        # Direct execution mode: print directly to terminal
        print(terminal_message, flush=True)

def _human_like_scroll_to_element(driver, element):
    """
    Scroll element into view with human-like behavior (smooth scroll)
    """
    try:
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        time.sleep(random.uniform(0.3, 0.6))  # Random wait after scroll
    except Exception as e:
        log(f"⚠️ Could not scroll to element: {e}", 'warning')
        try:
            # Fallback to instant scroll
            driver.execute_script("arguments[0].scrollIntoView(true);", element)
            time.sleep(0.3)
        except:
            pass

def _human_like_click(driver, element, wait_time_before=None, wait_time_after=None):
    """
    Perform human-like click using ActionChains (move to element, then click)
    """
    try:
        # Scroll element into view first
        _human_like_scroll_to_element(driver, element)
        
        # Wait before click (human-like delay)
        if wait_time_before is None:
            wait_time_before = random.uniform(0.2, 0.5)
        time.sleep(wait_time_before)
        
        # Use ActionChains to move mouse to element and click
        actions = ActionChains(driver)
        actions.move_to_element(element).pause(random.uniform(0.1, 0.2)).click().perform()
        
        # Wait after click (human-like delay)
        if wait_time_after is None:
            wait_time_after = random.uniform(0.2, 0.4)
        time.sleep(wait_time_after)
        
        return True
    except Exception as e:
        log(f"⚠️ Human-like click failed: {e}, trying fallback...", 'warning')
        try:
            # Fallback to JavaScript click if ActionChains fails
            driver.execute_script("arguments[0].click();", element)
            time.sleep(0.2)
            return True
        except Exception as e2:
            log(f"❌ Fallback click also failed: {e2}", 'error')
            raise

def _human_like_type(element, text, clear_first=True):
    """
    Type text with human-like delays between keystrokes
    """
    try:
        if clear_first:
            element.clear()
            time.sleep(random.uniform(0.1, 0.2))
        
        # Type with random delays between characters (simulating human typing)
        for char in text:
            element.send_keys(char)
            # Random delay between 0.05 and 0.15 seconds per character (human-like)
            time.sleep(random.uniform(0.05, 0.15))
        
        # Final delay after typing
        time.sleep(random.uniform(0.2, 0.4))
        return True
    except Exception as e:
        log(f"⚠️ Human-like typing failed: {e}, trying fallback...", 'warning')
        try:
            if clear_first:
                element.clear()
            element.send_keys(text)
            time.sleep(0.2)
            return True
        except Exception as e2:
            log(f"❌ Fallback typing also failed: {e2}", 'error')
            raise

def check_login_status_message(driver, wait=None):
    """Check and log the login status message from the page xpath: //*[@id="main"]/div/div[2]/div/div[1]/p
    
    Returns:
        tuple: (status_message, is_failure) where is_failure is True if message is "認証に失敗しました。"
    """
    try:
        if wait is None:
            wait = WebDriverWait(driver, 3)  # Short timeout for status check
        
        # Primary check: Try to find the login status message element at the specified xpath
        status_xpath = '//*[@id="main"]/div/div[2]/div/div[1]/p'
        try:
            # Use find_elements instead of wait.until to avoid exceptions if element doesn't exist
            status_elements = driver.find_elements(By.XPATH, status_xpath)
            if status_elements:
                status_element = status_elements[0]
                # Check if element is visible
                if status_element.is_displayed():
                    status_message = status_element.text.strip()
                    if status_message:
                        # Check for exact failure message
                        is_failure = status_message == "認証に失敗しました。"
                        # Determine log level based on message content
                        if is_failure or any(keyword in status_message for keyword in ['失敗', '失敗しました', 'エラー', 'error', 'fail']):
                            log(f"❌ Login status message (FAILURE): {status_message}", 'error')
                        elif any(keyword in status_message for keyword in ['成功', 'success', '完了', 'complete']):
                            log(f"✅ Login status message (SUCCESS): {status_message}", 'success')
                        else:
                            log(f"📋 Login status message: {status_message}", 'info')
                        return (status_message, is_failure)
                else:
                    # Element exists but is not visible - check text anyway
                    status_message = status_element.text.strip()
                    if status_message:
                        is_failure = status_message == "認証に失敗しました。"
                        log(f"📋 Login status message (hidden): {status_message}", 'info')
                        return (status_message, is_failure)
        except Exception as e:
            # Element not found or error accessing it - this is normal if login was successful
            pass
        
        # Secondary check: If we're still on login page, try to find any error messages in common locations
        if "login.html" in driver.current_url or "login-mfa.html" in driver.current_url:
            try:
                # Check for error messages in various possible locations
                error_selectors = [
                    '//*[@id="main"]//p[contains(@class, "error")]',
                    '//*[@id="main"]//div[contains(@class, "error")]',
                    '//*[@id="main"]//span[contains(@class, "error")]',
                    '//*[@id="main"]//*[contains(text(), "認証に失敗")]',
                    '//*[@id="main"]//*[contains(text(), "失敗")]',
                ]
                for selector in error_selectors:
                    try:
                        error_elements = driver.find_elements(By.XPATH, selector)
                        for elem in error_elements:
                            if elem.is_displayed():
                                error_text = elem.text.strip()
                                if error_text:
                                    is_failure = error_text == "認証に失敗しました。"
                                    log(f"⚠️ Error message found on page: {error_text}", 'warning')
                                    return (error_text, is_failure)
                    except:
                        continue
            except:
                pass
        
        return (None, False)
    except Exception as e:
        # Don't fail the login process if status check fails
        log(f"⚠️ Could not check login status message: {e}", 'warning')
        return (None, False)

def solve_recaptcha(site_key, url, max_retries=5):
    log(f"🔐 Starting CAPTCHA solving process... (Site key: {site_key[:20]}...)", 'info')
    for attempt in range(1, max_retries + 1):
        try:
            check_stop()  # Check if stopped before starting attempt
            log(f"🔄 Solving CAPTCHA... (Attempt {attempt}/{max_retries})", 'info')
            submit_url = f"http://2captcha.com/in.php?key={CAPTCHA_API_KEY}&method=userrecaptcha&googlekey={site_key}&pageurl={url}&invisible=1"
            response = requests.get(submit_url)
            
            if "OK|" not in response.text:
                error_msg = response.text
                log(f"❌ 2Captcha submit error: {error_msg}", 'error')
                if attempt < max_retries:
                    log(f"⏳ Retrying in 3 seconds...", 'warning')
                    for _ in range(3):
                        check_stop()  # Check stop during wait
                        time.sleep(1)
                    continue
                else:
                    raise Exception(f"2captcha error: {error_msg}")
            
            captcha_id = response.text.split("|")[1]
            log(f"📝 CAPTCHA submitted successfully. ID: {captcha_id}", 'info')
            
            result_url = f"http://2captcha.com/res.php?key={CAPTCHA_API_KEY}&action=get&id={captcha_id}"
            
            for i in range(30):
                check_stop()  # Check stop before each wait
                time.sleep(5)
                check_stop()  # Check stop after wait
                result = requests.get(result_url)
                
                if "CAPCHA_NOT_READY" in result.text:
                    log(f"⏳ Waiting for CAPTCHA solution... ({i+1}/30)", 'info')
                    continue
                elif "OK|" in result.text:
                    solution = result.text.split("|")[1]
                    log(f"✅ CAPTCHA solved successfully! (Solution length: {len(solution)} chars)", 'success')
                    return solution
                elif "ERROR_CAPTCHA_UNSOLVABLE" in result.text:
                    log(f"⚠️ CAPTCHA unsolvable, retrying... (Attempt {attempt}/{max_retries})", 'warning')
                    if attempt < max_retries:
                        for _ in range(3):
                            check_stop()
                            time.sleep(1)
                        break  # Break inner loop to retry from start
                    else:
                        raise Exception(f"2captcha error: {result.text}")
                else:
                    error_msg = result.text
                    log(f"❌ 2Captcha result error: {error_msg}", 'error')
                    if attempt < max_retries:
                        for _ in range(3):
                            check_stop()
                            time.sleep(1)
                        break  # Break inner loop to retry from start
                    else:
                        raise Exception(f"2captcha error: {error_msg}")
            else:
                # If we exhausted the 30 attempts without success, retry
                if attempt < max_retries:
                    log(f"⏱️ CAPTCHA timeout, retrying... (Attempt {attempt}/{max_retries})", 'warning')
                    for _ in range(3):
                        check_stop()
                        time.sleep(1)
                    continue
                else:
                    raise Exception("CAPTCHA timeout after all retries")
        except StopIteration:
            log("⏹️ CAPTCHA solving stopped by user", 'warning')
            raise
        except Exception as e:
            if attempt < max_retries and "2captcha error" in str(e):
                log(f"⚠️ Error occurred: {e}, retrying... (Attempt {attempt}/{max_retries})", 'warning')
                for _ in range(3):
                    check_stop()
                    time.sleep(1)
                continue
            else:
                raise
    
    raise Exception("CAPTCHA solving failed after all retries")

def get_otp_from_gmail():
    log("📧 Checking Gmail for OTP email...", 'info')
    check_stop()  # Check stop before starting
    service = get_service()
    
    # Use the current EMAIL variable instead of hardcoded email
    target_email = EMAIL.lower() if EMAIL else None
    if not target_email:
        raise ValueError("EMAIL is not set. Cannot retrieve OTP.")
    
    log(f"🔍 Looking for OTP emails sent to: {target_email}", 'info')
    
    for attempt in range(12):
        try:
            check_stop()  # Check stop before each attempt
            messages = list_messages(service, max_results=5, query='ポケモンセンターオンライン ログイン用パスコード')
            
            log(f"📬 Attempt {attempt+1}/12: Found {len(messages) if messages else 0} message(s) matching query", 'info')
            
            if messages:
                for msg in messages:
                    try:
                        msg_id = msg['id']
                        subject, snippet, sender, to, date, categories, body = get_message(service, msg_id)
                        
                        log(f"  📨 Checking email: Subject='{subject[:50]}...', To='{to}', Date='{date}'", 'info')
                        
                        # Check if email is sent to the current user's email
                        # Handle multiple recipients (to field can be a comma-separated list)
                        to_emails = [email.strip().lower() for email in to.split(',')] if to else []
                        
                        if target_email in to_emails or any(target_email in email for email in to_emails):
                            log(f"  ✅ Email recipient matches: {target_email}", 'success')
                            log(f"  📄 Body length: {len(body)} characters", 'info')
                            log(f"  👁️ Body preview: {body[:200]}...", 'info')
                            
                            # Try multiple regex patterns to find OTP
                            patterns = [
                                r'【パスコード】(\d{6})',  # Original pattern
                                r'パスコード[：:]\s*(\d{6})',  # Alternative format
                                r'認証コード[：:]\s*(\d{6})',  # Alternative format
                                r'コード[：:]\s*(\d{6})',  # Generic code
                                r'(\d{6})',  # Any 6-digit number (fallback)
                            ]
                            
                            for pattern in patterns:
                                match = re.search(pattern, body)
                                if match:
                                    otp = match.group(1)
                                    log(f"  ✅ OTP found with pattern '{pattern}': {otp}", 'success')
                                    return otp
                            
                            log(f"  ❌ No OTP pattern matched in email body", 'error')
                        else:
                            log(f"  ⚠️ Email recipient doesn't match. Expected: {target_email}, Got: {to}", 'warning')
                    except Exception as e:
                        log(f"  ❌ Error processing message {msg_id}: {e}", 'error')
                        import traceback
                        traceback.print_exc()
                        continue
            else:
                log(f"  📭 No messages found with query 'ポケモンセンターオンライン ログイン用パスコード'", 'info')
                # Try a broader search
                if attempt == 2:  # After 3 attempts, try broader search
                    log("  🔍 Trying broader search (searching for 'パスコード' or 'passcode')...", 'info')
                    broader_messages = list_messages(service, max_results=10, query='パスコード OR passcode')
                    log(f"  📬 Broader search found {len(broader_messages) if broader_messages else 0} message(s)", 'info')
        
        except StopIteration:
            log("⏹️ OTP retrieval stopped by user", 'warning')
            raise
        except Exception as e:
            log(f"  ❌ Error during Gmail search: {e}", 'error')
            import traceback
            traceback.print_exc()
        
        log(f"⏳ Waiting for OTP email... ({attempt+1}/12)", 'info')
        # Check stop during wait (split 5 seconds into 5 checks)
        for _ in range(5):
            check_stop()
            time.sleep(1)
    
    raise Exception("OTP not received after 12 attempts (60 seconds). Check if email was sent and verify email address matches.")
def _attempt_single_login(driver, wait, attempt_number=1):
    """
    Attempt a single login with email and password.
    Returns: (success, needs_retry) tuple where:
        - success: True if login was successful (redirected away from login page)
        - needs_retry: True if authentication failed and should retry
    """
    try:
        check_stop()  # Check stop before starting
        log(f"🌐 Opening login page: {LOGIN_URL} (Attempt {attempt_number})", 'info')
        driver.get(LOGIN_URL)
        log("⏳ Waiting for page to load...", 'info')
        # Check stop during page load wait
        for _ in range(10):
            check_stop()
            time.sleep(1)
        
        check_stop()  # Check stop before entering credentials
        log(f"📧 Entering email: {EMAIL}", 'info')
        email_field = wait.until(EC.presence_of_element_located((By.ID, "email")))
        _human_like_scroll_to_element(driver, email_field)
        _human_like_type(email_field, EMAIL)
        log("✅ Email entered successfully", 'success')
        check_stop()
        
        check_stop()
        log("🔒 Entering password...", 'info')
        if PASSWORD is None:
            raise ValueError("PASSWORD is not set. Please set it in .env file or include it in column B of the Excel file.")
        password_field = driver.find_element(By.ID, "password")
        _human_like_scroll_to_element(driver, password_field)
        _human_like_type(password_field, PASSWORD)
        log("✅ Password entered successfully", 'success')
        check_stop()
        
        check_stop()
        log("🖱️ Clicking login button...", 'info')
        login_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.loginBtn")))
        _human_like_click(driver, login_btn)
        log("⏳ Waiting for login response...", 'info')
        # Check stop during login wait
        for _ in range(8):
            check_stop()
            time.sleep(1)
        
        current_url = driver.current_url
        log(f"📍 Current URL after login attempt: {current_url}", 'info')
        
        # Check for authentication failure message
        log("🔍 Checking login status message from page...", 'info')
        status_message, is_failure = check_login_status_message(driver, wait)
        
        # Check if authentication failed
        if is_failure and status_message == "認証に失敗しました。":
            log(f"❌ Authentication failed: '{status_message}' (Attempt {attempt_number})", 'error')
            return (False, True)  # Failed, needs retry
        
        # If still on login page and failure message found, return needs retry
        if "login.html" in current_url and is_failure:
            log(f"⚠️ Still on login page with failure message (Attempt {attempt_number})", 'warning')
            return (False, True)  # Failed, needs retry
        
        # If redirected away from login page, login might be successful (but may need CAPTCHA/OTP)
        if "login.html" not in current_url:
            log(f"✅ Redirected away from login page (Attempt {attempt_number}) - login may be successful", 'success')
            return (True, False)  # Successful redirect, no retry needed
        
        # If still on login page but no failure message, may need CAPTCHA/OTP
        if "login.html" in current_url:
            if status_message:
                log(f"📋 Login status message: {status_message} (Attempt {attempt_number})", 'info')
            else:
                log("ℹ️ No login status message found, may need CAPTCHA/OTP", 'info')
            return (False, False)  # Not successful yet, but no authentication failure - may need CAPTCHA/OTP
        
        return (False, False)
    except StopIteration:
        raise
    except Exception as e:
        log(f"❌ Error during login attempt {attempt_number}: {e}", 'error')
        return (False, False)

def lottery_begin(driver, wait=None):
    if wait is None:
        wait = WebDriverWait(driver, 30)
    try:
        # Attempt login with retry mechanism (max 3 attempts)
        max_login_attempts = 3
        login_success = False
        needs_retry = False
        
        for attempt in range(1, max_login_attempts + 1):
            check_stop()
            log(f"🔐 Login attempt {attempt}/{max_login_attempts}...", 'info')
            
            success, needs_retry = _attempt_single_login(driver, wait, attempt_number=attempt)
            
            if success:
                log(f"✅ Login successful on attempt {attempt}!", 'success')
                login_success = True
                break
            
            # Check if authentication failed with "認証に失敗しました。" message
            if needs_retry:
                if attempt < max_login_attempts:
                    log(f"⚠️ Authentication failed on attempt {attempt}. Retrying... (Attempt {attempt + 1}/{max_login_attempts})", 'warning')
                    time.sleep(2)  # Wait before retry
                    continue
                else:
                    # Last attempt failed, raise exception to move to next email
                    log(f"❌ Authentication failed after {max_login_attempts} attempts. Proceeding to next login...", 'error')
                    raise Exception(f"Authentication failed after {max_login_attempts} attempts for email: {EMAIL}")
            else:
                # No authentication failure message detected, but may need CAPTCHA/OTP - continue with normal flow
                log("ℹ️ No authentication failure message detected. Continuing with CAPTCHA/OTP flow if needed...", 'info')
                break
        
        # If authentication failed after all attempts (shouldn't reach here if exception was raised)
        if not login_success and needs_retry:
            raise Exception(f"Authentication failed after {max_login_attempts} attempts for email: {EMAIL}")
        
        current_url = driver.current_url
        
        check_stop()  # Check stop before CAPTCHA
        if "login.html" in current_url and "認証に失敗" in driver.page_source:
            log("⚠️ Login failed - CAPTCHA required. Starting CAPTCHA solving...", 'warning')
            
            match = re.search(r'6Le[a-zA-Z0-9_-]+', driver.page_source)
            if match:
                site_key = match.group(0)
                log(f"🔑 Found reCAPTCHA site key: {site_key[:30]}...", 'info')
                
                try:
                    captcha_solution = solve_recaptcha(site_key, driver.current_url)
                except StopIteration:
                    log("⏹️ Login process stopped during CAPTCHA solving", 'warning')
                    raise
                
                log("💉 Injecting CAPTCHA solution into page...", 'info')
                driver.execute_script(f'''
                    var callback = function(token) {{
                        var textareas = document.getElementsByName("g-recaptcha-response");
                        for (var i = 0; i < textareas.length; i++) {{
                            textareas[i].value = token;
                        }}
                    }};
                    callback("{captcha_solution}");
                    
                    if (typeof ___grecaptcha_cfg !== 'undefined') {{
                        Object.keys(___grecaptcha_cfg.clients).forEach(function(key) {{
                            var client = ___grecaptcha_cfg.clients[key];
                            if (client && client.callback) {{
                                client.callback("{captcha_solution}");
                            }}
                        }});
                    }}
                ''')
                log("✅ CAPTCHA solution injected", 'success')
                time.sleep(2)
                
                log("🔄 Re-entering credentials after CAPTCHA...", 'info')
                email_field = driver.find_element(By.ID, "email")
                _human_like_scroll_to_element(driver, email_field)
                _human_like_type(email_field, EMAIL)
                log("✅ Email re-entered", 'success')
                
                password_field = driver.find_element(By.ID, "password")
                _human_like_scroll_to_element(driver, password_field)
                _human_like_type(password_field, PASSWORD)
                log("✅ Password re-entered", 'success')
                
                log("🖱️ Clicking login button again...", 'info')
                login_btn = driver.find_element(By.CSS_SELECTOR, "a.loginBtn")
                _human_like_click(driver, login_btn)
                log("⏳ Waiting for login response after CAPTCHA...", 'info')
                # Check stop during login wait
                for _ in range(8):
                    check_stop()
                    time.sleep(1)
                
                # Check login status message after CAPTCHA retry
                log("🔍 Checking login status message after CAPTCHA submission...", 'info')
                status_message, is_failure = check_login_status_message(driver, wait)
                if status_message:
                    if is_failure:
                        log(f"❌ Login status message after CAPTCHA (FAILURE): {status_message}", 'error')
                        # Check if we should retry (but we've already tried once, so check attempt count)
                        # If this is within the retry loop, it will be handled by the retry mechanism
                    else:
                        log(f"📋 Login status message after CAPTCHA: {status_message}", 'info')
                else:
                    log("ℹ️ No login status message found after CAPTCHA", 'info')
                
                # Check current URL after CAPTCHA
                current_url_after_captcha = driver.current_url
                if "login.html" in current_url_after_captcha and is_failure and status_message == "認証に失敗しました。":
                    log("❌ Authentication still failed after CAPTCHA. Will retry if attempts remaining...", 'error')
                    # This will be caught by the retry loop if attempts remain
        
        check_stop()  # Check stop before OTP
        if "login-mfa" in driver.current_url or "パスコード" in driver.page_source:
            log("🔐 OTP (One-Time Password) required for login", 'info')
            
            log("⏳ Waiting for OTP email to be sent (5 seconds)...", 'info')
            # Check stop during OTP wait
            for _ in range(5):
                check_stop()
                time.sleep(1)
            
            try:
                otp = get_otp_from_gmail()
                log(f"✅ OTP retrieved from Gmail: {otp}", 'success')
            except StopIteration:
                log("⏹️ Login process stopped during OTP retrieval", 'warning')
                raise
            
            check_stop()
            log("⌨️ Entering OTP into form...", 'info')
            otp_field = wait.until(EC.presence_of_element_located((By.ID, "authCode")))
            _human_like_scroll_to_element(driver, otp_field)
            _human_like_type(otp_field, otp)
            log("✅ OTP entered successfully", 'success')
            check_stop()
            
            check_stop()
            log("🖱️ Submitting OTP...", 'info')
            submit_btn = wait.until(EC.element_to_be_clickable((By.ID, "certify")))
            _human_like_click(driver, submit_btn)
            log("⏳ OTP submitted, waiting for authentication response...", 'info')
            # Check stop during OTP response wait
            for _ in range(10):
                check_stop()
                time.sleep(1)
            
            log(f"📍 Current URL after OTP submission: {driver.current_url}", 'info')
            
            # Check login status message after OTP submission
            log("🔍 Checking login status message after OTP submission...", 'info')
            status_message, is_failure = check_login_status_message(driver, wait)
            if status_message:
                if is_failure:
                    log(f"❌ Login status message after OTP (FAILURE): {status_message}", 'error')
                else:
                    log(f"📋 Login status message after OTP: {status_message}", 'info')
            else:
                log("ℹ️ No login status message found after OTP", 'info')
            
            # Check if authentication failed after OTP
            if is_failure and status_message == "認証に失敗しました。":
                log("❌ Authentication failed after OTP submission. Maximum retries reached for OTP flow.", 'error')
                raise Exception(f"Authentication failed after OTP for email: {EMAIL}")
            
            if "login-mfa" in driver.current_url and "認証に失敗" in driver.page_source:
                log("⚠️ OTP authentication failed - retrieving fresh OTP and retrying...", 'warning')
                for _ in range(3):
                    check_stop()
                    time.sleep(1)
                
                try:
                    otp = get_otp_from_gmail()
                    log(f"✅ Fresh OTP retrieved: {otp}", 'success')
                except StopIteration:
                    log("⏹️ Login process stopped during OTP retry", 'warning')
                    raise
                
                check_stop()
                log(f"⌨️ Entering fresh OTP: {otp}", 'info')
                otp_field = driver.find_element(By.ID, "authCode")
                _human_like_scroll_to_element(driver, otp_field)
                _human_like_type(otp_field, otp)
                log("✅ Fresh OTP entered", 'success')
                check_stop()
                
                check_stop()
                log("🖱️ Submitting fresh OTP...", 'info')
                submit_btn = driver.find_element(By.ID, "certify")
                _human_like_click(driver, submit_btn)
                log("⏳ Waiting for authentication response after retry...", 'info')
                # Check stop during retry wait
                for _ in range(10):
                    check_stop()
                    time.sleep(1)
                
                log(f"📍 Current URL after OTP retry: {driver.current_url}", 'info')
                
                # Check login status message after OTP retry
                log("🔍 Checking login status message after OTP retry...", 'info')
                status_message, is_failure = check_login_status_message(driver, wait)
                if status_message:
                    if is_failure:
                        log(f"❌ Login status message after OTP retry (FAILURE): {status_message}", 'error')
                    else:
                        log(f"📋 Login status message after OTP retry: {status_message}", 'info')
                else:
                    log("ℹ️ No login status message found after OTP retry", 'info')
                
                # Check if authentication still failed after OTP retry
                if is_failure and status_message == "認証に失敗しました。":
                    log("❌ Authentication failed after OTP retry. Maximum retries reached.", 'error')
                    raise Exception(f"Authentication failed after OTP retry for email: {EMAIL}")
        
        log(f"📍 Final URL after login process: {driver.current_url}", 'info')
        
        # Final login status check
        log("🔍 Performing final login status message check...", 'info')
        status_message, is_failure = check_login_status_message(driver, wait)
        if status_message:
            if is_failure:
                log(f"❌ Final login status message (FAILURE): {status_message}", 'error')
                # If we still have authentication failure at this point, it means retries didn't work
                raise Exception(f"Authentication failed after all attempts for email: {EMAIL}")
            else:
                log(f"📋 Final login status message: {status_message}", 'info')
        else:
            log("ℹ️ No login status message found in final check", 'info')
        
        # Navigate to apply page if not already there
        if "apply.html" not in driver.current_url:
            log(f"🌐 Navigating to application page: {APPLY_URL}", 'info')
            driver.get(APPLY_URL)
            # Wait for page to fully load
            for _ in range(5):
                check_stop()
                time.sleep(1)
        else:
            log("✅ Already on application page", 'success')
            # Wait a moment for page to stabilize if already on apply page
            check_stop()
            time.sleep(2)
        
        # Wait for page to be fully ready and check for any initial pop04 (normal information modal after successful login)
        log("⏳ Waiting for apply page to fully load and checking for pop04 modal...", 'info')
        check_stop()
        time.sleep(3)  # Give page time to load and show any pop04
        
        pop04_reloaded = False  # Flag to track if page was reloaded due to exception
        max_reload_attempts = 5  # Maximum number of reload attempts (changed from 6 to 5)
        reload_attempt = 0
        
        # Check for pop04 with exception message "意図しない例外が発生しました。" and reload up to 5 times
        # If pop04 appears without exception message (normal case after successful login), just close it
        while reload_attempt < max_reload_attempts:
            check_stop()
            reload_attempt += 1
            
            exception_detected = False
            try:
                pop04_elements = driver.find_elements(By.ID, "pop04")
                if pop04_elements and pop04_elements[0].is_displayed():
                    pop04_element = pop04_elements[0]
                    log("🔔 Popup (pop04) detected, checking content...", 'info')
                    
                    # Check if pop04 contains "意図しない例外が発生しました。" message
                    pop04_message_xpath = '//*[@id="pop04"]/div/div[1]/p'  # Define xpath outside try block for reuse
                    try:
                        pop04_message_element = driver.find_element(By.XPATH, pop04_message_xpath)
                        pop04_message_text = pop04_message_element.text.strip()
                        log(f"📋 Pop04 message: {pop04_message_text}", 'info')
                        
                        if "意図しない例外が発生しました。" in pop04_message_text:
                            exception_detected = True
                            log(f"⚠️ Exception message detected in pop04 (Attempt {reload_attempt}/{max_reload_attempts}): '意図しない例外が発生しました。' - Reloading page...", 'warning')
                            
                            # Reload the page when exception message is detected
                            check_stop()
                            reload_success = False
                            
                            # Store current URL before reload for verification
                            current_url_before = driver.current_url
                            log(f"📍 Current URL before reload: {current_url_before}", 'info')
                            
                            # Try multiple reload methods in order of reliability
                            try:
                                # Method 1: Use driver.refresh() (most reliable for Selenium)
                                log(f"🔄 Attempt {reload_attempt}/{max_reload_attempts}: Using driver.refresh() to reload page...", 'info')
                                driver.refresh()
                                check_stop()
                                # Wait for page to load and verify reload
                                for wait_iteration in range(5):
                                    check_stop()
                                    time.sleep(1)
                                    try:
                                        # Check if page has reloaded by verifying driver state
                                        driver.current_url
                                        break
                                    except:
                                        continue
                                
                                current_url_after = driver.current_url
                                log(f"📍 Current URL after refresh: {current_url_after}", 'info')
                                log(f"✅ Page reloaded via refresh() (Attempt {reload_attempt})", 'success')
                                reload_success = True
                                pop04_reloaded = True
                            except Exception as e:
                                log(f"⚠️ Could not reload via refresh(): {e}. Trying driver.get() method...", 'warning')
                                try:
                                    # Method 2: Use driver.get() with current URL
                                    log(f"🔄 Attempt {reload_attempt}/{max_reload_attempts}: Using driver.get() method...", 'info')
                                    current_url = driver.current_url
                                    if not current_url:
                                        current_url = APPLY_URL  # Fallback to apply URL
                                    driver.get(current_url)
                                    check_stop()
                                    # Wait for page to load
                                    for wait_iteration in range(5):
                                        check_stop()
                                        time.sleep(1)
                                    log(f"✅ Page reloaded via get() (Attempt {reload_attempt})", 'success')
                                    reload_success = True
                                    pop04_reloaded = True
                                except Exception as e2:
                                    log(f"⚠️ Could not reload via get(): {e2}. Trying JavaScript location.reload()...", 'warning')
                                    try:
                                        # Method 3: Use JavaScript location.reload() - more reliable than F5
                                        log(f"🔄 Attempt {reload_attempt}/{max_reload_attempts}: Using JavaScript location.reload()...", 'info')
                                        driver.execute_script("window.location.reload(true);")  # Force reload from server (bypass cache)
                                        check_stop()
                                        # Wait for page to reload
                                        for wait_iteration in range(5):
                                            check_stop()
                                            time.sleep(1)
                                        log(f"✅ Page reloaded via JavaScript location.reload() (Attempt {reload_attempt})", 'success')
                                        reload_success = True
                                        pop04_reloaded = True
                                    except Exception as e3:
                                        log(f"⚠️ Could not reload via JavaScript: {e3}. Trying alternative JavaScript method...", 'warning')
                                        try:
                                            # Method 4: Alternative JavaScript method
                                            log(f"🔄 Attempt {reload_attempt}/{max_reload_attempts}: Using JavaScript window.location = window.location...", 'info')
                                            driver.execute_script("window.location = window.location;")
                                            check_stop()
                                            # Wait for page to reload
                                            for wait_iteration in range(5):
                                                check_stop()
                                                time.sleep(1)
                                            log(f"✅ Page reloaded via alternative JavaScript method (Attempt {reload_attempt})", 'success')
                                            reload_success = True
                                            pop04_reloaded = True
                                        except Exception as e4:
                                            log(f"❌ All reload methods failed (Attempt {reload_attempt}). Last error: {e4}", 'error')
                                            time.sleep(2)
                                            reload_success = False
                            
                            if reload_success:
                                # Wait for page to stabilize after reload
                                check_stop()
                                log(f"⏳ Waiting for page to stabilize after reload {reload_attempt}...", 'info')
                                time.sleep(3)
                                
                                # Check if pop04 appears again after reload
                                check_stop()
                                try:
                                    pop04_check = driver.find_elements(By.ID, "pop04")
                                    if pop04_check and pop04_check[0].is_displayed():
                                        # Check if exception message is still present
                                        try:
                                            pop04_message_check = driver.find_element(By.XPATH, pop04_message_xpath)
                                            pop04_message_text_check = pop04_message_check.text.strip()
                                            if "意図しない例外が発生しました。" in pop04_message_text_check:
                                                if reload_attempt < max_reload_attempts:
                                                    log(f"⚠️ Pop04 with exception message still present after reload {reload_attempt}. Will retry...", 'warning')
                                                    continue  # Continue loop to retry reload
                                                else:
                                                    log(f"❌ Pop04 with exception message still present after {max_reload_attempts} reload attempts. Proceeding anyway...", 'error')
                                                    # Try to close pop04 and proceed
                                                    try:
                                                        pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                                        pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                                        _human_like_click(driver, pop04_link)
                                                        log("✅ Pop04 modal closed after max reload attempts", 'success')
                                                        time.sleep(1)
                                                    except:
                                                        pass
                                                    break  # Exit reload loop
                                            else:
                                                log("✅ Exception message cleared after reload. Pop04 may have different message now.", 'success')
                                                # Try to close pop04 normally
                                                try:
                                                    pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                                    pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                                    _human_like_click(driver, pop04_link)
                                                    log("✅ Pop04 modal closed", 'success')
                                                except:
                                                    pass
                                                break  # Exit reload loop - exception cleared
                                        except:
                                            # Exception message element not found, pop04 might have different content
                                            log("✅ Exception message not found after reload. Pop04 may have been cleared or changed.", 'success')
                                            try:
                                                pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                                pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                                _human_like_click(driver, pop04_link)
                                                log("✅ Pop04 modal closed", 'success')
                                                time.sleep(1)
                                            except:
                                                pass
                                            break  # Exit reload loop
                                    else:
                                        log(f"✅ Pop04 cleared after reload {reload_attempt}. No pop04 detected.", 'success')
                                        break  # Exit reload loop - pop04 cleared
                                except Exception as e:
                                    log(f"⚠️ Error checking pop04 after reload {reload_attempt}: {e}. Assuming cleared.", 'warning')
                                    break  # Exit reload loop - assume cleared
                            else:
                                # Reload failed, but continue to next attempt if attempts remain
                                if reload_attempt < max_reload_attempts:
                                    log(f"⚠️ Reload attempt {reload_attempt} failed. Will retry...", 'warning')
                                    time.sleep(2)  # Wait before retry
                                    continue
                                else:
                                    log(f"❌ All {max_reload_attempts} reload attempts failed. Proceeding anyway...", 'error')
                                    # Try to close pop04 and proceed
                                    try:
                                        pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                        pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                        _human_like_click(driver, pop04_link)
                                        log("✅ Pop04 modal closed after all reload attempts failed", 'success')
                                    except:
                                        pass
                                    break  # Exit reload loop
                        else:
                            # Normal pop04 handling (not exception case) - this is expected after successful OTP login
                            log("ℹ️ Pop04 detected but no exception message. This is normal after successful login. Closing pop04...", 'info')
                            try:
                                # Wait a moment for pop04 to be fully ready
                                time.sleep(1)
                                pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                # Try multiple methods to close pop04
                                pop04_closed = False
                                
                                # Method 1: Try to find and click the link element
                                try:
                                    pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)), timeout=5)
                                    _human_like_click(driver, pop04_link)
                                    log("✅ Pop04 closed via link click", 'success')
                                    pop04_closed = True
                                except Exception as e1:
                                    log(f"⚠️ Could not close pop04 via link: {e1}. Trying alternative methods...", 'warning')
                                    # Method 2: Try to find and click any close button or link in pop04
                                    try:
                                        close_buttons = driver.find_elements(By.XPATH, '//*[@id="pop04"]//a | //*[@id="pop04"]//button')
                                        for btn in close_buttons:
                                            if btn.is_displayed() and btn.is_enabled():
                                                _human_like_click(driver, btn)
                                                log("✅ Pop04 closed via alternative button", 'success')
                                                pop04_closed = True
                                                break
                                    except Exception as e2:
                                        log(f"⚠️ Could not close pop04 via alternative buttons: {e2}", 'warning')
                                
                                if pop04_closed:
                                    # Verify pop04 is closed
                                    time.sleep(1)
                                    try:
                                        pop04_verify = driver.find_elements(By.ID, "pop04")
                                        if not pop04_verify or not pop04_verify[0].is_displayed():
                                            log("✅ Pop04 successfully closed and verified", 'success')
                                        else:
                                            log("⚠️ Pop04 still appears to be open after close attempt. Proceeding anyway...", 'warning')
                                    except:
                                        log("✅ Pop04 close verification completed", 'success')
                                else:
                                    log("⚠️ Could not close pop04 using any method. It may close automatically or may not affect functionality. Proceeding...", 'warning')
                                
                            except Exception as e:
                                log(f"⚠️ Error handling normal pop04: {e}. Proceeding as it's not an error case...", 'warning')
                            break  # Exit reload loop - not exception case, normal pop04 handled
                    except Exception as e:
                        log(f"⚠️ Could not read pop04 message (Attempt {reload_attempt}): {e}. Treating as normal pop04 and closing...", 'warning')
                        try:
                            # Try to close pop04 as normal modal (not exception case)
                            time.sleep(1)
                            pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                            pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)), timeout=5)
                            _human_like_click(driver, pop04_link)
                            log("✅ Pop04 closed (fallback method)", 'success')
                        except Exception as e2:
                            log(f"⚠️ Could not close pop04 (fallback): {e2}. It may close automatically. Proceeding...", 'warning')
                        break  # Exit reload loop - treat as normal case
                else:
                    # pop04 not displayed, exit loop - this is normal after successful login
                    log(f"ℹ️ Pop04 not displayed (Attempt {reload_attempt}). This is normal. Ready for lottery processing.", 'info')
                    break
            except Exception as e:
                # pop04 not found or error checking - this is normal if pop04 doesn't appear
                log(f"ℹ️ Pop04 not present or error checking (Attempt {reload_attempt}): {e}. This is normal. Ready for lottery processing.", 'info')
                break  # Exit reload loop - treat as normal case
        
        # Final verification: ensure pop04 is cleared before proceeding to lottery processing
        check_stop()
        log("🔍 Final check: Verifying pop04 is cleared before starting lottery processing...", 'info')
        try:
            # Wait a moment for any pop04 animations to complete
            time.sleep(2)
            pop04_final_check = driver.find_elements(By.ID, "pop04")
            if pop04_final_check and pop04_final_check[0].is_displayed():
                log("⚠️ Pop04 still present before lottery processing. Attempting final close...", 'warning')
                try:
                    # Try multiple methods to close pop04
                    pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                    pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)), timeout=5)
                    _human_like_click(driver, pop04_link)
                    log("✅ Pop04 closed in final check", 'success')
                    time.sleep(random.uniform(1.5, 2.5))  # Wait for pop04 to fully close
                    
                    # Verify pop04 is closed
                    pop04_verify_check = driver.find_elements(By.ID, "pop04")
                    if not pop04_verify_check or not pop04_verify_check[0].is_displayed():
                        log("✅ Pop04 successfully closed and verified in final check", 'success')
                    else:
                        log("⚠️ Pop04 may still be open, but proceeding with lottery processing...", 'warning')
                except Exception as e:
                    log(f"⚠️ Could not close pop04 in final check: {e}. Proceeding with lottery processing anyway...", 'warning')
            else:
                log("✅ Pop04 is not present (or already closed). Ready for lottery processing.", 'success')
        except Exception as e:
            log(f"⚠️ Error in final pop04 check: {e}. Proceeding with lottery processing...", 'warning')
        
        # If page was reloaded due to exception, wait for page to stabilize
        if pop04_reloaded:
            check_stop()
            log(f"✅ Page reload completed due to exception. Total reload attempts: {reload_attempt}. Waiting for page to stabilize...", 'success')
            time.sleep(3)  # Wait for page to fully load after reload
        
        # Ensure we're on the apply page before starting lottery processing
        check_stop()
        if "apply.html" not in driver.current_url:
            log(f"⚠️ Not on apply page before lottery processing. Navigating to apply page...", 'warning')
            try:
                driver.get(APPLY_URL)
                log(f"✅ Navigated to apply page", 'success')
                # Wait for page to load
                for _ in range(5):
                    check_stop()
                    time.sleep(1)
            except Exception as e:
                log(f"⚠️ Could not navigate to apply page: {e}. Continuing anyway...", 'warning')
        
        # Normal flow: Process all available lotteries sequentially
        log(f"🎰 Starting lottery entry process for up to {_max_lotteries} lotteries...", 'info')
        check_stop()
        
        lottery_result = _process_all_lotteries(driver, wait, max_lotteries=_max_lotteries)
        
        log("🎉 All lottery entry processes completed!", 'success')
        return lottery_result

    except StopIteration:
        log("⏹️ Login process stopped by user", 'warning')
        # Return failure result when stopped by user
        return {
            'results': [],
            'final_status': '中断',
            'message': 'ユーザーによって中断されました'
        }
    except Exception as e:
        log(f"❌ Fatal error in login process: {e}", 'error')
        import traceback
        traceback.print_exc()
        # Return failure result when error occurs
        return {
            'results': [],
            'final_status': '失敗',
            'message': f'ログインエラー: {str(e)[:100]}'
        }

def _check_lottery_status(driver, wait, lottery_number):
    """
    Check the status of a specific lottery number.
    Returns: (status_text, exists) tuple where status_text is the status and exists is True if lottery exists
    """
    try:
        # First, ensure we're on the apply page
        if "apply.html" not in driver.current_url:
            log(f"⚠️ Not on apply page (current URL: {driver.current_url}). Navigating to apply page...", 'warning')
            try:
                driver.get(APPLY_URL)
                for _ in range(3):
                    check_stop()
                    time.sleep(1)
            except Exception as e:
                log(f"⚠️ Could not navigate to apply page: {e}", 'warning')
                return (None, False)
        
        # Check if the lottery list container exists
        main_list_xpath = '//*[@id="main"]/div[1]/ul'
        try:
            main_list = driver.find_elements(By.XPATH, main_list_xpath)
            if not main_list or len(main_list) == 0:
                log(f"⚠️ Lottery list container not found. Page might not be loaded correctly.", 'warning')
                return (None, False)
        except Exception as e:
            log(f"⚠️ Could not find lottery list container: {e}", 'warning')
            return (None, False)
        
        # Check if this specific lottery exists
        status_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/div/span[1]'
        status_elements = driver.find_elements(By.XPATH, status_xpath)
        
        if status_elements and len(status_elements) > 0:
            status_element = status_elements[0]
            # Try to get text even if element is not displayed (might be in collapsed state)
            try:
                status_text = status_element.text.strip()
                if status_text:
                    return (status_text, True)
                else:
                    # Element exists but has no text - might be hidden, try to get it anyway
                    status_text = driver.execute_script("return arguments[0].textContent || arguments[0].innerText || '';", status_element)
                    if status_text and status_text.strip():
                        return (status_text.strip(), True)
                    return (None, True)  # Element exists but status is empty
            except Exception as e:
                log(f"⚠️ Could not get status text for lottery #{lottery_number}: {e}", 'warning')
                return (None, True)  # Element exists but couldn't read text
        else:
            return (None, False)  # Element does not exist
    except Exception as e:
        log(f"⚠️ Could not check status for lottery #{lottery_number}: {e}", 'warning')
        return (None, False)

def _check_and_solve_captcha_on_apply_page(driver, wait):
    """
    Check for CAPTCHA on the apply.html page and solve it if present.
    Returns True if CAPTCHA was found and solved, False otherwise.
    """
    try:
        check_stop()
        log("🔍 Checking for CAPTCHA on apply page...", 'info')
        
        # Check if we're on apply page
        if "apply.html" not in driver.current_url:
            log("ℹ️ Not on apply page, skipping CAPTCHA check", 'info')
            return False
        
        # Look for reCAPTCHA in the page source
        page_source = driver.page_source
        match = re.search(r'6Le[a-zA-Z0-9_-]+', page_source)
        
        if match:
            site_key = match.group(0)
            log(f"🔑 Found reCAPTCHA on apply page. Site key: {site_key[:30]}...", 'info')
            
            try:
                # Solve CAPTCHA using 2captcha API
                check_stop()
                captcha_solution = solve_recaptcha(site_key, driver.current_url)
                
                log("💉 Injecting CAPTCHA solution into apply page...", 'info')
                # Inject CAPTCHA solution into the page
                driver.execute_script(f'''
                    var callback = function(token) {{
                        var textareas = document.getElementsByName("g-recaptcha-response");
                        for (var i = 0; i < textareas.length; i++) {{
                            textareas[i].value = token;
                        }}
                    }};
                    callback("{captcha_solution}");
                    
                    if (typeof ___grecaptcha_cfg !== 'undefined') {{
                        Object.keys(___grecaptcha_cfg.clients).forEach(function(key) {{
                            var client = ___grecaptcha_cfg.clients[key];
                            if (client && client.callback) {{
                                client.callback("{captcha_solution}");
                            }}
                        }});
                    }}
                ''')
                log("✅ CAPTCHA solution injected into apply page", 'success')
                time.sleep(2)  # Wait for CAPTCHA to be processed
                check_stop()
                return True
            except StopIteration:
                log("⏹️ CAPTCHA solving stopped by user", 'warning')
                raise
            except Exception as e:
                log(f"❌ Error solving CAPTCHA on apply page: {e}", 'error')
                return False
        else:
            log("ℹ️ No CAPTCHA found on apply page", 'info')
            return False
    except Exception as e:
        log(f"⚠️ Error checking for CAPTCHA on apply page: {e}", 'warning')
        return False

def _process_all_lotteries(driver, wait, max_lotteries=1):
    """
    Process all available lotteries sequentially.
    - Checks up to max_lotteries number of lotteries
    - Skips lotteries with status "受付終了" (already closed)
    - Skips lotteries with status "受付完了" (already completed)
    - Processes only lotteries with status "受付中" (currently open)
    - Continues checking even if a lottery doesn't exist (up to max_lotteries)
    - Continues to next lottery even if one fails
    - If pop04/pop05 error occurs, reloads page and restarts from first lottery
    - Tracks completed lotteries and skips them on restart
    
    Returns:
        dict: {
            'results': [{'lottery': int, 'status': str, 'reason': str}, ...],
            'final_status': str,  # '成功' or '失敗'
            'message': str  # Detailed message for Excel column D
        }
    """
    # Track completed lotteries (successfully processed or already completed)
    completed_lotteries = set()  # Set of lottery numbers that are completed
    lottery_results = []  # Track results for each lottery
    max_retry_attempts = 10  # Maximum number of retry attempts (to prevent infinite loop)
    retry_attempt = 0
    
    # Initialize final_status and final_message to avoid UnboundLocalError
    final_status = '失敗'  # Default status
    final_message = '処理中にエラーが発生しました'
    
    while retry_attempt < max_retry_attempts:
        if retry_attempt == 0:
            log(f"🔍 Starting to check up to {max_lotteries} lotteries for processing...", 'info')
        else:
            log(f"🔄 Retry attempt {retry_attempt + 1}/{max_retry_attempts}: Restarting from first lottery after reload...", 'info')
            log(f"📋 Completed lotteries so far: {sorted(completed_lotteries) if completed_lotteries else 'None'}", 'info')
        
        reload_occurred = False  # Track if reload occurred during processing
        lottery_number = 1
        checked_count = 0  # Number of lotteries checked in this attempt
        
        # Check for CAPTCHA on apply page before starting lottery processing
        check_stop()
        captcha_solved = _check_and_solve_captcha_on_apply_page(driver, wait)
        if captcha_solved:
            log("✅ CAPTCHA solved before starting lottery processing", 'success')
            time.sleep(2)  # Wait for page to stabilize after CAPTCHA solution
        
        # Ensure we're on the apply page at the start
        check_stop()
        if "apply.html" not in driver.current_url:
            log(f"⚠️ Not on apply page at start. Navigating to apply page...", 'warning')
            try:
                driver.get(APPLY_URL)
                log(f"✅ Navigated to apply page", 'success')
                # Wait for page to load
                for _ in range(5):
                    check_stop()
                    time.sleep(1)
            except Exception as e:
                log(f"⚠️ Could not navigate to apply page: {e}", 'warning')
        
        # Continue checking until we've checked max_lotteries number of lotteries
        while checked_count < max_lotteries:
            check_stop()
            
            # Ensure we're on the apply page before checking lottery status
            if "apply.html" not in driver.current_url:
                log(f"⚠️ Not on apply page before checking lottery #{lottery_number}. Navigating to apply page...", 'warning')
                try:
                    driver.get(APPLY_URL)
                    log(f"✅ Navigated to apply page", 'success')
                    # Wait for page to load
                    for _ in range(5):
                        check_stop()
                        time.sleep(1)
                except Exception as e:
                    log(f"⚠️ Could not navigate to apply page: {e}. Continuing anyway...", 'warning')
            
            # Skip if this lottery was already completed
            if lottery_number in completed_lotteries:
                log(f"⏭️ Lottery #{lottery_number} already completed. Skipping to next lottery...", 'info')
                lottery_number += 1
                checked_count += 1
                continue
            
            # Check if this lottery exists and get its status
            status_text, exists = _check_lottery_status(driver, wait, lottery_number)
            checked_count += 1  # Count this lottery as checked
            
            if not exists:
                log(f"📋 Lottery #{lottery_number} does not exist. Continuing to check next lottery ({checked_count}/{max_lotteries} checked)...", 'info')
                lottery_results.append({
                    'lottery': lottery_number,
                    'status': '存在しない',
                    'reason': f'抽選{lottery_number}は存在しません'
                })
                lottery_number += 1
                continue  # Skip to next lottery number, but continue checking
            
            # Log status if available
            if status_text:
                log(f"📊 Lottery #{lottery_number} status: '{status_text}'", 'info')
            else:
                log(f"📊 Lottery #{lottery_number} exists but status is empty or unavailable. Skipping...", 'warning')
                lottery_number += 1
                continue
            
            # Skip if already closed (受付終了)
            if status_text == "受付終了":
                log(f"⏭️ Lottery #{lottery_number} is already closed (受付終了). Skipping to next lottery...", 'warning')
                lottery_results.append({
                    'lottery': lottery_number,
                    'status': 'スキップ(終了)',
                    'reason': f'抽選{lottery_number}は受付終了しています'
                })
                lottery_number += 1
                continue
            
            # Skip if already completed (受付完了)
            if status_text == "受付完了":
                log(f"⏭️ Lottery #{lottery_number} is already completed (受付完了). Marking as completed and skipping...", 'warning')
                completed_lotteries.add(lottery_number)  # Mark as completed
                lottery_results.append({
                    'lottery': lottery_number,
                    'status': 'スキップ(完了)',
                    'reason': f'抽選{lottery_number}は受付完了しています'
                })
                lottery_number += 1
                continue
            
            # Process only if open (受付中)
            if status_text == "受付中":
                log(f"✅ Lottery #{lottery_number} is available (status: '{status_text}'). Processing...", 'success')
                check_stop()
                time.sleep(1)  # Brief wait before processing
                
                try:
                    result = _process_lottery_entry(driver, wait, lottery_number)
                    
                    # Check if reload is needed (pop04/pop05 error detected after apply button click)
                    if result == 'reload_needed':
                        # Page was reloaded due to exception - restart from first lottery
                        log("⚠️ Page reloaded due to pop04/pop05 exception after apply button click. Restarting from first lottery...", 'warning')
                        log(f"📋 Completed lotteries: {sorted(completed_lotteries)}. Will skip these on restart.", 'info')
                        reload_occurred = True
                        retry_attempt += 1
                        break  # Exit inner loop to restart from first lottery
                    
                    if result == True:
                        log(f"✅ Lottery #{lottery_number} processed successfully!", 'success')
                        # Mark this lottery as completed
                        completed_lotteries.add(lottery_number)
                        lottery_results.append({
                            'lottery': lottery_number,
                            'status': '成功',
                            'reason': f'抽選{lottery_number}の処理が成功しました'
                        })
                        
                        # Check for pop04/pop05 exception message after successful lottery processing
                        check_stop()
                        pop_reload_needed = _check_and_handle_pop_exceptions(driver, wait)
                        
                        if pop_reload_needed:
                            # Page was reloaded due to exception - restart from first lottery
                            log("⚠️ Page reloaded due to pop04/pop05 exception. Restarting from first lottery...", 'warning')
                            log(f"📋 Completed lotteries: {sorted(completed_lotteries)}. Will skip these on restart.", 'info')
                            reload_occurred = True
                            retry_attempt += 1
                            break  # Exit inner loop to restart from first lottery
                        else:
                            # No reload needed, move to next lottery
                            check_stop()
                            time.sleep(1)
                            lottery_number += 1
                    else:
                        log(f"⚠️ Lottery #{lottery_number} processing failed. Continuing to next lottery...", 'warning')
                        lottery_results.append({
                            'lottery': lottery_number,
                            'status': '失敗',
                            'reason': f'抽選{lottery_number}の処理が失敗しました'
                        })
                        # Check for pop04/pop05 exception message after failed lottery processing
                        check_stop()
                        pop_reload_needed = _check_and_handle_pop_exceptions(driver, wait)
                        
                        if pop_reload_needed:
                            # Page was reloaded due to exception - restart from first lottery
                            log("⚠️ Page reloaded due to pop04/pop05 exception. Restarting from first lottery...", 'warning')
                            log(f"📋 Completed lotteries: {sorted(completed_lotteries)}. Will skip these on restart.", 'info')
                            reload_occurred = True
                            retry_attempt += 1
                            break  # Exit inner loop to restart from first lottery
                        else:
                            # No reload needed, move to next lottery
                            if checked_count < max_lotteries:
                                check_stop()
                                try:
                                    if "apply.html" not in driver.current_url:
                                        log(f"🔄 Navigating back to apply page for next lottery check...", 'info')
                                        driver.get(APPLY_URL)
                                        for _ in range(3):
                                            check_stop()
                                            time.sleep(1)
                                except Exception as e:
                                    log(f"⚠️ Could not navigate to apply page: {e}. Continuing anyway...", 'warning')
                            lottery_number += 1
                except StopIteration:
                    log(f"⏹️ Lottery processing stopped by user at lottery #{lottery_number}", 'warning')
                    lottery_results.append({
                        'lottery': lottery_number,
                        'status': '中断',
                        'reason': f'抽選{lottery_number}の処理がユーザーによって中断されました'
                    })
                    raise
                except Exception as e:
                    log(f"❌ Error processing lottery #{lottery_number}: {e}. Continuing to next lottery...", 'error')
                    lottery_results.append({
                        'lottery': lottery_number,
                        'status': '失敗',
                        'reason': f'抽選{lottery_number}の処理でエラーが発生しました: {str(e)[:100]}'
                    })
                    # Check for pop04/pop05 exception message after error
                    check_stop()
                    pop_reload_needed = _check_and_handle_pop_exceptions(driver, wait)
                    
                    if pop_reload_needed:
                        # Page was reloaded due to exception - restart from first lottery
                        log("⚠️ Page reloaded due to pop04/pop05 exception. Restarting from first lottery...", 'warning')
                        log(f"📋 Completed lotteries: {sorted(completed_lotteries)}. Will skip these on restart.", 'info')
                        reload_occurred = True
                        retry_attempt += 1
                        break  # Exit inner loop to restart from first lottery
                    else:
                        # No reload needed, move to next lottery
                        if checked_count < max_lotteries:
                            check_stop()
                            try:
                                if "apply.html" not in driver.current_url:
                                    log(f"🔄 Navigating back to apply page after error for next lottery check...", 'info')
                                    driver.get(APPLY_URL)
                                    for _ in range(3):
                                        check_stop()
                                        time.sleep(1)
                            except Exception as e2:
                                log(f"⚠️ Could not navigate to apply page after error: {e2}. Continuing anyway...", 'warning')
                        lottery_number += 1
            else:
                log(f"⚠️ Lottery #{lottery_number} has unexpected status: '{status_text}'. Skipping...", 'warning')
                lottery_results.append({
                    'lottery': lottery_number,
                    'status': '不明',
                    'reason': f'抽選{lottery_number}のステータスが不明です: {status_text}'
                })
                lottery_number += 1
        
        # Check if reload occurred - if yes, restart from first lottery
        if reload_occurred:
            log("🔄 Reload occurred. Will restart from first lottery on next attempt...", 'info')
            continue  # Continue to next retry attempt
        
        # Check if all lotteries have been checked
        if checked_count >= max_lotteries:
            # Count completed lotteries (success + skipped completed)
            processed_count = len([r for r in lottery_results if r['status'] == '成功'])
            skipped_completed_count = len([r for r in lottery_results if r['status'] == 'スキップ(完了)'])
            skipped_closed_count = len([r for r in lottery_results if r['status'] == 'スキップ(終了)'])
            failed_count = len([r for r in lottery_results if r['status'] == '失敗'])
            
            log(f"✅ Completed checking {checked_count} lotteries (max: {max_lotteries})", 'info')
            log(f"📊 Lottery processing summary: {processed_count} processed, {skipped_completed_count} skipped (completed), {skipped_closed_count} skipped (closed), {failed_count} failed", 'info')
            log(f"📋 Completed lotteries: {sorted(completed_lotteries)}", 'info')
            
            # After all lotteries have been checked, verify if there are any "受付中" lotteries remaining
            log("🔍 Verifying if there are any remaining '受付中' lotteries...", 'info')
            has_open_lotteries = False
            for check_lottery_num in range(1, max_lotteries + 1):
                check_stop()
                status_text, exists = _check_lottery_status(driver, wait, check_lottery_num)
                if exists and status_text == "受付中":
                    log(f"⚠️ Lottery #{check_lottery_num} is still '受付中'. Will retry processing...", 'warning')
                    has_open_lotteries = True
                    # Remove from completed_lotteries if it was marked as completed
                    if check_lottery_num in completed_lotteries:
                        completed_lotteries.remove(check_lottery_num)
                        # Update result status
                        for result in lottery_results:
                            if result['lottery'] == check_lottery_num and result['status'] == '成功':
                                result['status'] = '失敗'
                                result['reason'] = f'抽選{check_lottery_num}は再確認時に「受付中」でした'
                                break
            
            # If there are open lotteries, retry processing
            if has_open_lotteries:
                log("🔄 Found remaining '受付中' lotteries. Starting retry attempt to process them...", 'info')
                retry_attempt += 1
                continue  # Continue to next retry attempt
            
            # Check if all required lotteries are completed (no open lotteries found)
            if len(completed_lotteries) >= max_lotteries:
                log(f"🎉 All {max_lotteries} lotteries have been completed and verified!", 'success')
                # Final verification: check all lotteries one more time to ensure they are all completed
                all_verified_completed = True
                for verify_lottery_num in range(1, max_lotteries + 1):
                    check_stop()
                    status_text, exists = _check_lottery_status(driver, wait, verify_lottery_num)
                    if exists and status_text not in ["受付完了", "受付終了"]:
                        if status_text == "受付中":
                            log(f"⚠️ Final verification: Lottery #{verify_lottery_num} is still '受付中'. Not all lotteries completed.", 'warning')
                            all_verified_completed = False
                            break
                
                if all_verified_completed:
                    final_status = '成功'
                    final_message = '成功'
                    log(f"✅ Final verification passed: All lotteries are completed!", 'success')
                    break  # Exit retry loop - all lotteries completed and verified
                else:
                    # Some lotteries are still open, continue to retry
                    log("🔄 Final verification failed: Some lotteries are still open. Starting another retry attempt...", 'info')
                    retry_attempt += 1
                    continue
            else:
                # Not all lotteries completed, but no reload occurred - proceed to final status determination
                # Continue to final status determination code below
                pass
        
        # Determine final status and create message for Excel column D
        # (Only if we've checked all lotteries)
        if checked_count >= max_lotteries:
            # Rules:
            # 1. All lotteries succeeded → "成功"
            # 2. All lotteries skipped (completed) → "成功"
            # 3. Any failure or skipped (closed) → "失敗: 詳細"
            # 4. Mixed results (some success, some failure/skipped closed) → "失敗: 詳細"
            
            has_failure = False
            has_skipped_closed = False
            has_skipped_completed = False
            has_success = False
            has_not_exist = False
            has_interrupted = False
            
            detail_parts = []
            
            detail_parts = []
            
            for result in lottery_results:
                status = result['status']
                lottery_num = result['lottery']
                
                if status == '成功':
                    has_success = True
                    detail_parts.append(f'抽選{lottery_num}成功')
                elif status == '失敗':
                    has_failure = True
                    detail_parts.append(f'抽選{lottery_num}失敗')
                elif status == 'スキップ(終了)':
                    has_skipped_closed = True
                    detail_parts.append(f'抽選{lottery_num}受付終了')
                elif status == 'スキップ(完了)':
                    has_skipped_completed = True
                    detail_parts.append(f'抽選{lottery_num}受付完了')
                elif status == '存在しない':
                    has_not_exist = True
                    detail_parts.append(f'抽選{lottery_num}存在しない')
                elif status == '中断':
                    has_interrupted = True
                    detail_parts.append(f'抽選{lottery_num}中断')
                elif status == '不明':
                    detail_parts.append(f'抽選{lottery_num}不明')
            
            # Determine final status according to requirements:
            # 1. All lotteries succeeded → "成功"
            # 2. All lotteries skipped (completed) → "成功"
            # 3. Any failure or skipped (closed) or not exist → "失敗"
            # 4. Mixed (some success, some failure) → "失敗"
            # 5. Mixed (some success, some skipped completed) → Check: if all are success or skipped completed, it's success
            
            # Log detailed information for debugging
            log(f"🔍 Analyzing {len(lottery_results)} lottery results for final status...", 'info')
            log(f"🔍 Results breakdown: success={has_success}, skipped_completed={has_skipped_completed}, skipped_closed={has_skipped_closed}, failure={has_failure}, not_exist={has_not_exist}, interrupted={has_interrupted}", 'info')
            log(f"🔍 Detail parts: {detail_parts}", 'info')
            
            if has_interrupted:
                final_status = '中断'
                final_message = '中断: ' + '、'.join(detail_parts)
                log(f"📋 Final status determined: {final_status} (interrupted)", 'info')
            elif has_failure or has_skipped_closed or has_not_exist:
                # If there's any failure, skipped (closed), or not exist, it's a failure
                final_status = '失敗'
                final_message = '失敗: ' + '、'.join(detail_parts)
                log(f"📋 Final status determined: {final_status} (has failure/skipped_closed/not_exist)", 'info')
            else:
                # Check if all lotteries are either success or skipped (completed)
                # This covers both cases: all success, all skipped (completed), or mixed success + skipped (completed)
                all_success_or_completed = True
                log(f"🔍 Checking if all lotteries are success or skipped (completed)...", 'info')
                for result in lottery_results:
                    status = result['status']
                    lottery_num = result['lottery']
                    log(f"🔍 Checking lottery {lottery_num}: status = '{status}'", 'info')
                    if status not in ['成功', 'スキップ(完了)']:
                        log(f"🔍 Lottery {lottery_num} status '{status}' is not success or skipped (completed). All success/completed check failed.", 'info')
                        all_success_or_completed = False
                        break
                
                if all_success_or_completed:
                    # All lotteries are success or skipped (completed), but need final verification
                    final_status = '成功'
                    final_message = '成功'
                    log(f"📋 Preliminary status determined: {final_status} (all lotteries are success or skipped completed)", 'success')
                    # Continue to final verification below
                else:
                    # Shouldn't reach here due to previous checks, but handle it
                    final_status = '失敗'
                    final_message = '失敗: ' + '、'.join(detail_parts)
                    log(f"📋 Final status determined: {final_status} (unexpected case - not all success/completed)", 'warning')
                    # Break from retry loop after determining final status
                    break
            
            # Perform final verification before breaking
            if final_status == '成功':
                # Final verification: check all lotteries one more time to ensure they are all completed
                log("🔍 Performing final verification: Checking all lotteries one more time...", 'info')
                all_final_verified = True
                for final_verify_num in range(1, max_lotteries + 1):
                    check_stop()
                    status_text, exists = _check_lottery_status(driver, wait, final_verify_num)
                    if exists:
                        if status_text == "受付中":
                            log(f"⚠️ Final verification: Lottery #{final_verify_num} is still '受付中'. Not all lotteries completed.", 'warning')
                            all_final_verified = False
                            # Remove from completed if it was marked as completed
                            if final_verify_num in completed_lotteries:
                                completed_lotteries.remove(final_verify_num)
                            # Update result
                            for result in lottery_results:
                                if result['lottery'] == final_verify_num:
                                    if result['status'] == '成功':
                                        result['status'] = '失敗'
                                        result['reason'] = f'抽選{final_verify_num}は最終確認時に「受付中」でした'
                                    break
                
                if all_final_verified:
                    log(f"✅ Final verification passed: All lotteries are completed. Final status: {final_status}", 'success')
                    break  # Exit retry loop
                else:
                    # Some lotteries are still open - retry if attempts remain
                    if retry_attempt < max_retry_attempts - 1:
                        log("🔄 Final verification failed: Some lotteries are still open. Starting another retry attempt...", 'info')
                        retry_attempt += 1
                        continue
                    else:
                        # Max attempts reached - use failure status
                        final_status = '失敗'
                        final_message = '失敗: 最終確認時に「受付中」の抽選が残っていました'
                        log(f"⚠️ Max retry attempts reached. Final status: {final_status}", 'warning')
                        break
            else:
                # Final status is already failure - break
                break
    
    # After retry loop completes - determine final status
    log(f"📋 Final lottery result: {final_status} - {final_message}", 'info')
    
    return {
        'results': lottery_results,
        'final_status': final_status,
        'message': final_message
    }

def _check_and_handle_pop_exceptions(driver, wait, max_reload_attempts=5):
    """
    Check for pop04 or pop05 with exception messages and reload if found.
    - pop04: "意図しない例外が発生しました。"
    - pop05: "一定時間操作していなかったため、OKボタンタップして再度開けてください。"
    Returns True if page was reloaded, False otherwise.
    """
    try:
        # First check pop05
        pop05_elements = driver.find_elements(By.ID, "pop05")
        if pop05_elements and pop05_elements[0].is_displayed():
            pop05_element = pop05_elements[0]
            log("🔔 Popup (pop05) detected, checking content...", 'info')
            
            # Check if pop05 contains timeout message
            pop05_message_xpath = '//*[@id="pop05"]/div/div[1]/p'
            try:
                pop05_message_element = driver.find_element(By.XPATH, pop05_message_xpath)
                pop05_message_text = pop05_message_element.text.strip()
                log(f"📋 Pop05 message: {pop05_message_text}", 'info')
                
                if "一定時間操作していなかったため、OKボタンタップして再度開けてください。" in pop05_message_text:
                    log(f"⚠️ Timeout message detected in pop05: '一定時間操作していなかったため、OKボタンタップして再度開けてください。' - Reloading page...", 'warning')
                    
                    # Reload the page when timeout message is detected (up to max_reload_attempts times)
                    reload_attempt = 0
                    while reload_attempt < max_reload_attempts:
                        check_stop()
                        reload_attempt += 1
                        log(f"🔄 Reload attempt {reload_attempt}/{max_reload_attempts} (pop05)...", 'info')
                        
                        reload_success = False
                        current_url_before = driver.current_url
                        log(f"📍 Current URL before reload: {current_url_before}", 'info')
                        
                        # Try multiple reload methods
                        try:
                            # Method 1: driver.refresh()
                            driver.refresh()
                            check_stop()
                            for _ in range(5):
                                check_stop()
                                time.sleep(1)
                            log(f"✅ Page reloaded via refresh() (Attempt {reload_attempt})", 'success')
                            reload_success = True
                        except Exception as e:
                            log(f"⚠️ Could not reload via refresh(): {e}. Trying driver.get()...", 'warning')
                            try:
                                # Method 2: driver.get()
                                current_url = driver.current_url or APPLY_URL
                                driver.get(current_url)
                                check_stop()
                                for _ in range(5):
                                    check_stop()
                                    time.sleep(1)
                                log(f"✅ Page reloaded via get() (Attempt {reload_attempt})", 'success')
                                reload_success = True
                            except Exception as e2:
                                log(f"⚠️ Could not reload via get(): {e2}. Trying JavaScript...", 'warning')
                                try:
                                    # Method 3: JavaScript location.reload()
                                    driver.execute_script("window.location.reload(true);")
                                    check_stop()
                                    for _ in range(5):
                                        check_stop()
                                        time.sleep(1)
                                    log(f"✅ Page reloaded via JavaScript (Attempt {reload_attempt})", 'success')
                                    reload_success = True
                                except Exception as e3:
                                    log(f"❌ All reload methods failed (Attempt {reload_attempt}): {e3}", 'error')
                                    reload_success = False
                        
                        if reload_success:
                            # Wait for page to stabilize
                            log(f"⏳ Waiting for page to stabilize after reload {reload_attempt}...", 'info')
                            time.sleep(3)
                            
                            # Check if pop05 still exists with timeout message
                            check_stop()
                            try:
                                pop05_check = driver.find_elements(By.ID, "pop05")
                                if pop05_check and pop05_check[0].is_displayed():
                                    try:
                                        pop05_message_check = driver.find_element(By.XPATH, pop05_message_xpath)
                                        pop05_message_text_check = pop05_message_check.text.strip()
                                        if "一定時間操作していなかったため、OKボタンタップして再度開けてください。" in pop05_message_text_check:
                                            if reload_attempt < max_reload_attempts:
                                                log(f"⚠️ Timeout message still present after reload {reload_attempt}. Retrying...", 'warning')
                                                continue  # Retry reload
                                            else:
                                                log(f"❌ Timeout message still present after {max_reload_attempts} reload attempts. Closing pop05...", 'error')
                                                # Try to close pop05
                                                try:
                                                    pop05_link_xpath = '//*[@id="pop05"]/div/div[1]/ul/li/a'
                                                    pop05_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop05_link_xpath)))
                                                    _human_like_click(driver, pop05_link)
                                                    log("✅ Pop05 closed after max reload attempts", 'success')
                                                    time.sleep(1)
                                                except:
                                                    pass
                                                return True  # Page was reloaded
                                        else:
                                            log("✅ Timeout message cleared after reload. Closing pop05...", 'success')
                                            # Close pop05 normally
                                            try:
                                                pop05_link_xpath = '//*[@id="pop05"]/div/div[1]/ul/li/a'
                                                pop05_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop05_link_xpath)))
                                                _human_like_click(driver, pop05_link)
                                                log("✅ Pop05 closed", 'success')
                                                time.sleep(1)
                                            except:
                                                pass
                                            return True  # Page was reloaded
                                    except:
                                        log("✅ Pop05 message check completed. Closing pop05...", 'success')
                                        # Try to close pop05
                                        try:
                                            pop05_link_xpath = '//*[@id="pop05"]/div/div[1]/ul/li/a'
                                            pop05_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop05_link_xpath)))
                                            _human_like_click(driver, pop05_link)
                                            log("✅ Pop05 closed", 'success')
                                            time.sleep(1)
                                        except:
                                            pass
                                        return True  # Page was reloaded
                                else:
                                    log(f"✅ Pop05 cleared after reload {reload_attempt}", 'success')
                                    return True  # Page was reloaded
                            except Exception as e:
                                log(f"⚠️ Error checking pop05 after reload: {e}. Assuming cleared.", 'warning')
                                return True  # Page was reloaded
                        else:
                            # Reload failed
                            if reload_attempt < max_reload_attempts:
                                log(f"⚠️ Reload attempt {reload_attempt} failed. Will retry...", 'warning')
                                time.sleep(2)
                                continue
                            else:
                                log(f"❌ All {max_reload_attempts} reload attempts failed. Closing pop05...", 'error')
                                # Try to close pop05
                                try:
                                    pop05_link_xpath = '//*[@id="pop05"]/div/div[1]/ul/li/a'
                                    pop05_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop05_link_xpath)))
                                    _human_like_click(driver, pop05_link)
                                    log("✅ Pop05 closed after all reload attempts failed", 'success')
                                    time.sleep(1)
                                except:
                                    pass
                                return False  # Page reload failed, but tried
                else:
                    # Pop05 exists but no timeout message - close it normally
                    log("ℹ️ Pop05 detected but no timeout message. Closing normally...", 'info')
                    try:
                        pop05_link_xpath = '//*[@id="pop05"]/div/div[1]/ul/li/a'
                        pop05_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop05_link_xpath)))
                        _human_like_click(driver, pop05_link)
                        log("✅ Pop05 closed", 'success')
                        time.sleep(1)
                    except Exception as e:
                        log(f"⚠️ Could not close pop05: {e}", 'warning')
                    return False  # No reload needed
            except Exception as e:
                log(f"⚠️ Could not read pop05 message: {e}. Trying to close pop05...", 'warning')
                try:
                    pop05_link_xpath = '//*[@id="pop05"]/div/div[1]/ul/li/a'
                    pop05_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop05_link_xpath)))
                    _human_like_click(driver, pop05_link)
                    log("✅ Pop05 closed (fallback)", 'success')
                    time.sleep(1)
                except:
                    pass
                return False  # No reload needed
        
        # Then check pop04 (original logic)
        pop04_elements = driver.find_elements(By.ID, "pop04")
        check_stop()
        pop04_elements = driver.find_elements(By.ID, "pop04")
        if pop04_elements and pop04_elements[0].is_displayed():
            pop04_element = pop04_elements[0]
            log("🔔 Popup (pop04) detected during lottery processing, checking content...", 'info')
            
            # Check if pop04 contains "意図しない例外が発生しました。" message
            pop04_message_xpath = '//*[@id="pop04"]/div/div[1]/p'
            try:
                pop04_message_element = driver.find_element(By.XPATH, pop04_message_xpath)
                pop04_message_text = pop04_message_element.text.strip()
                log(f"📋 Pop04 message: {pop04_message_text}", 'info')
                
                if "意図しない例外が発生しました。" in pop04_message_text:
                    log(f"⚠️ Exception message detected in pop04: '意図しない例外が発生しました。' - Reloading page...", 'warning')
                    
                    # Reload the page when exception message is detected (up to max_reload_attempts times)
                    reload_attempt = 0
                    while reload_attempt < max_reload_attempts:
                        check_stop()
                        reload_attempt += 1
                        log(f"🔄 Reload attempt {reload_attempt}/{max_reload_attempts}...", 'info')
                        
                        reload_success = False
                        current_url_before = driver.current_url
                        log(f"📍 Current URL before reload: {current_url_before}", 'info')
                        
                        # Try multiple reload methods
                        try:
                            # Method 1: driver.refresh()
                            driver.refresh()
                            check_stop()
                            for _ in range(5):
                                check_stop()
                                time.sleep(1)
                            log(f"✅ Page reloaded via refresh() (Attempt {reload_attempt})", 'success')
                            reload_success = True
                        except Exception as e:
                            log(f"⚠️ Could not reload via refresh(): {e}. Trying driver.get()...", 'warning')
                            try:
                                # Method 2: driver.get()
                                current_url = driver.current_url or APPLY_URL
                                driver.get(current_url)
                                check_stop()
                                for _ in range(5):
                                    check_stop()
                                    time.sleep(1)
                                log(f"✅ Page reloaded via get() (Attempt {reload_attempt})", 'success')
                                reload_success = True
                            except Exception as e2:
                                log(f"⚠️ Could not reload via get(): {e2}. Trying JavaScript...", 'warning')
                                try:
                                    # Method 3: JavaScript location.reload()
                                    driver.execute_script("window.location.reload(true);")
                                    check_stop()
                                    for _ in range(5):
                                        check_stop()
                                        time.sleep(1)
                                    log(f"✅ Page reloaded via JavaScript (Attempt {reload_attempt})", 'success')
                                    reload_success = True
                                except Exception as e3:
                                    log(f"❌ All reload methods failed (Attempt {reload_attempt}): {e3}", 'error')
                                    reload_success = False
                        
                        if reload_success:
                            # Wait for page to stabilize
                            log(f"⏳ Waiting for page to stabilize after reload {reload_attempt}...", 'info')
                            time.sleep(3)
                            
                            # Check if pop04 still exists with exception message
                            check_stop()
                            try:
                                pop04_check = driver.find_elements(By.ID, "pop04")
                                if pop04_check and pop04_check[0].is_displayed():
                                    try:
                                        pop04_message_check = driver.find_element(By.XPATH, pop04_message_xpath)
                                        pop04_message_text_check = pop04_message_check.text.strip()
                                        if "意図しない例外が発生しました。" in pop04_message_text_check:
                                            if reload_attempt < max_reload_attempts:
                                                log(f"⚠️ Exception message still present after reload {reload_attempt}. Retrying...", 'warning')
                                                continue  # Retry reload
                                            else:
                                                log(f"❌ Exception message still present after {max_reload_attempts} reload attempts. Closing pop04...", 'error')
                                                # Try to close pop04
                                                try:
                                                    pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                                    pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                                    _human_like_click(driver, pop04_link)
                                                    log("✅ Pop04 closed after max reload attempts", 'success')
                                                    time.sleep(1)
                                                except:
                                                    pass
                                                return True  # Page was reloaded
                                        else:
                                            log("✅ Exception message cleared after reload. Closing pop04...", 'success')
                                            # Close pop04 normally
                                            try:
                                                pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                                pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                                _human_like_click(driver, pop04_link)
                                                log("✅ Pop04 closed", 'success')
                                                time.sleep(1)
                                            except:
                                                pass
                                            return True  # Page was reloaded
                                    except:
                                        log("✅ Pop04 message check completed. Closing pop04...", 'success')
                                        # Try to close pop04
                                        try:
                                            pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                            pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                            _human_like_click(driver, pop04_link)
                                            log("✅ Pop04 closed", 'success')
                                            time.sleep(1)
                                        except:
                                            pass
                                        return True  # Page was reloaded
                                else:
                                    log(f"✅ Pop04 cleared after reload {reload_attempt}", 'success')
                                    return True  # Page was reloaded
                            except Exception as e:
                                log(f"⚠️ Error checking pop04 after reload: {e}. Assuming cleared.", 'warning')
                                return True  # Page was reloaded
                        else:
                            # Reload failed
                            if reload_attempt < max_reload_attempts:
                                log(f"⚠️ Reload attempt {reload_attempt} failed. Will retry...", 'warning')
                                time.sleep(2)
                                continue
                            else:
                                log(f"❌ All {max_reload_attempts} reload attempts failed. Closing pop04...", 'error')
                                # Try to close pop04
                                try:
                                    pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                                    pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                                    driver.execute_script("arguments[0].click();", pop04_link)
                                    log("✅ Pop04 closed after all reload attempts failed", 'success')
                                    time.sleep(1)
                                except:
                                    pass
                                return False  # Page reload failed, but tried
                else:
                    # Pop04 exists but no exception message - close it normally
                    log("ℹ️ Pop04 detected but no exception message. Closing normally...", 'info')
                    try:
                        pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                        pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                        driver.execute_script("arguments[0].click();", pop04_link)
                        log("✅ Pop04 closed", 'success')
                        time.sleep(1)
                    except Exception as e:
                        log(f"⚠️ Could not close pop04: {e}", 'warning')
                    return False  # No reload needed
            except Exception as e:
                log(f"⚠️ Could not read pop04 message: {e}. Trying to close pop04...", 'warning')
                try:
                    pop04_link_xpath = '//*[@id="pop04"]/div/div[1]/ul/li/a'
                    pop04_link = wait.until(EC.element_to_be_clickable((By.XPATH, pop04_link_xpath)))
                    driver.execute_script("arguments[0].click();", pop04_link)
                    log("✅ Pop04 closed (fallback)", 'success')
                    time.sleep(1)
                except:
                    pass
                return False  # No reload needed
        else:
            # No pop04 detected
            return False  # No reload needed
    except Exception as e:
        log(f"⚠️ Error checking pop04: {e}. Assuming no action needed.", 'warning')
        return False  # No reload needed

def _process_lottery_entry(driver, wait, lottery_number=1):
    """
    Process lottery entry for a specific lottery number.
    Returns:
        - True on success
        - False on failure
        - 'reload_needed' if pop04/pop05 error detected and page was reloaded (needs restart from first lottery)
    """
    try:
        check_stop()
        log(f"🎰 Processing lottery #{lottery_number}...", 'info')
        
        # Check for CAPTCHA on apply page before processing lottery entry
        check_stop()
        captcha_solved = _check_and_solve_captcha_on_apply_page(driver, wait)
        if captcha_solved:
            log(f"✅ CAPTCHA solved before processing lottery #{lottery_number}", 'success')
            time.sleep(1)  # Brief wait after CAPTCHA solution
        
        # Step 2: Click on dt element to expand details
        check_stop()
        log(f"🖱️ Clicking lottery #{lottery_number} details (dt)...", 'info')
        dt_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dt'
        dt_element = wait.until(EC.element_to_be_clickable((By.XPATH, dt_xpath)))
        _human_like_click(driver, dt_element)
        check_stop()
        time.sleep(random.uniform(0.5, 1.0))  # Wait for details to expand
        
        # Step 3: Click radio button - try multiple strategies to ensure it works
        check_stop()
        log(f"☑️ Selecting radio button for lottery #{lottery_number}...", 'info')
        
        radio_clicked = False
        
        # Strategy 1: Find and click p.radio element (most reliable - clicking p will trigger label/input)
        try:
            check_stop()
            log(f"  🔍 Strategy 1: Looking for p.radio element...", 'info')
            p_radio_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form/ul[1]/li/p[@class="radio"]'
            p_radio_element = wait.until(EC.presence_of_element_located((By.XPATH, p_radio_xpath)))
            
            if p_radio_element.is_displayed():
                # Find the input inside p.radio
                input_elem = p_radio_element.find_element(By.TAG_NAME, 'input')
                # Set checked and trigger events
                driver.execute_script("""
                    arguments[0].checked = true;
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('click', { bubbles: true }));
                """, input_elem)
                # Also click the label to ensure it's properly selected
                try:
                    label_elem = p_radio_element.find_element(By.TAG_NAME, 'label')
                    driver.execute_script("arguments[0].click();", label_elem)
                except:
                    pass
                
                # Verify it's checked
                is_checked = driver.execute_script("return arguments[0].checked;", input_elem)
                if is_checked:
                    log(f"✅ Radio button selected successfully via p.radio element for lottery #{lottery_number}", 'success')
                    radio_clicked = True
        except Exception as e:
            log(f"  ⚠️ Strategy 1 failed: {str(e)[:80]}...", 'warning')
        
        # Strategy 2: Find and click label element (label click automatically selects input)
        if not radio_clicked:
            try:
                check_stop()
                log(f"  🔍 Strategy 2: Looking for label element...", 'info')
                label_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form/ul[1]/li/p[@class="radio"]/label'
                label_element = wait.until(EC.element_to_be_clickable((By.XPATH, label_xpath)))
                
                driver.execute_script("arguments[0].click();", label_element)
                
                # Verify input is checked
                input_elem = label_element.find_element(By.TAG_NAME, 'input')
                is_checked = driver.execute_script("return arguments[0].checked;", input_elem)
                if is_checked:
                    log(f"✅ Radio button selected successfully via label element for lottery #{lottery_number}", 'success')
                    radio_clicked = True
            except Exception as e:
                log(f"  ⚠️ Strategy 2 failed: {str(e)[:80]}...", 'warning')
        
        # Strategy 3: Find and click input[type="radio"] directly
        if not radio_clicked:
            try:
                check_stop()
                log(f"  🔍 Strategy 3: Looking for input[type=\"radio\"] element...", 'info')
                input_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form/ul[1]/li/p[@class="radio"]/label/input[@type="radio"]'
                input_element = wait.until(EC.presence_of_element_located((By.XPATH, input_xpath)))
                
                # Set checked property and trigger events
                driver.execute_script("""
                    arguments[0].checked = true;
                    arguments[0].click();
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """, input_element)
                
                # Verify it's checked
                is_checked = driver.execute_script("return arguments[0].checked;", input_element)
                if is_checked:
                    log(f"✅ Radio button selected successfully via input element for lottery #{lottery_number}", 'success')
                    radio_clicked = True
            except Exception as e:
                log(f"  ⚠️ Strategy 3 failed: {str(e)[:80]}...", 'warning')
        
        # Strategy 4: Fallback - try span element
        if not radio_clicked:
            try:
                check_stop()
                log(f"  🔍 Strategy 4: Looking for span element (fallback)...", 'info')
                span_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form/ul[1]/li/p[@class="radio"]/label/span'
                span_element = wait.until(EC.element_to_be_clickable((By.XPATH, span_xpath)))
                
                driver.execute_script("arguments[0].click();", span_element)
                
                # Find and verify input is checked
                p_parent = span_element.find_element(By.XPATH, './ancestor::p[@class="radio"]')
                input_elem = p_parent.find_element(By.TAG_NAME, 'input')
                is_checked = driver.execute_script("return arguments[0].checked;", input_elem)
                if is_checked:
                    log(f"✅ Radio button selected successfully via span element for lottery #{lottery_number}", 'success')
                    radio_clicked = True
            except Exception as e:
                log(f"  ⚠️ Strategy 4 failed: {str(e)[:80]}...", 'warning')
        
        # Strategy 5: Last resort - find first radio input in form
        if not radio_clicked:
            try:
                check_stop()
                log(f"  🔍 Strategy 5: Looking for first radio input in form (last resort)...", 'info')
                form_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form'
                form_element = driver.find_element(By.XPATH, form_xpath)
                input_element = form_element.find_element(By.XPATH, './/input[@type="radio"]')
                
                driver.execute_script("""
                    arguments[0].checked = true;
                    arguments[0].click();
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """, input_element)
                
                is_checked = driver.execute_script("return arguments[0].checked;", input_element)
                if is_checked:
                    log(f"✅ Radio button selected successfully via fallback method for lottery #{lottery_number}", 'success')
                    radio_clicked = True
            except Exception as e:
                log(f"  ⚠️ Strategy 5 failed: {str(e)[:80]}...", 'warning')
        
        if not radio_clicked:
            log(f"❌ Could not select radio button using any strategy for lottery #{lottery_number}", 'error')
            raise Exception(f"Failed to select radio button for lottery #{lottery_number} after trying all strategies")
        
        check_stop()
        time.sleep(1)
        
        # Step 4: Check checkbox
        check_stop()
        log(f"✅ Checking checkbox for lottery #{lottery_number}...", 'info')
        checkbox_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form/div/div'
        checkbox_element = wait.until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath)))
        _human_like_scroll_to_element(driver, checkbox_element)
        time.sleep(random.uniform(0.2, 0.4))
        _human_like_click(driver, checkbox_element)
        log(f"✅ Checkbox checked for lottery #{lottery_number}", 'success')
        check_stop()
        
        # Step 5: Check for CAPTCHA before submitting application
        check_stop()
        log(f"🔍 Checking for CAPTCHA before submitting lottery #{lottery_number}...", 'info')
        captcha_solved_before_submit = _check_and_solve_captcha_on_apply_page(driver, wait)
        if captcha_solved_before_submit:
            log(f"✅ CAPTCHA solved before submitting lottery #{lottery_number}", 'success')
            time.sleep(1)  # Brief wait after CAPTCHA solution
        
        # Step 6: Click submit button to open modal
        check_stop()
        log(f"🔔 Clicking submit button for lottery #{lottery_number} to open modal...", 'info')
        submit_xpath = f'//*[@id="main"]/div[1]/ul/li[{lottery_number}]/div[2]/dl/dd/div[3]/form/ul[2]/li/a'
        submit_element = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
        _human_like_scroll_to_element(driver, submit_element)
        time.sleep(random.uniform(0.3, 0.5))
        _human_like_click(driver, submit_element)
        check_stop()
        time.sleep(random.uniform(1.5, 2.5))  # Wait for modal to appear
        
        # Step 7: Wait for modal to appear and click apply button
        check_stop()
        log(f"🎯 Waiting for modal (pop01) to appear for lottery #{lottery_number}...", 'info')
        modal_xpath = '//*[@id="pop01"]/div/div[1]'
        try:
            modal_element = wait.until(EC.presence_of_element_located((By.XPATH, modal_xpath)))
            log("✅ Modal appeared", 'success')
        except Exception as e:
            log(f"⚠️ Modal element not found with xpath {modal_xpath}: {e}. Trying to proceed anyway...", 'warning')
            # Try alternative: just wait for applyBtn
            pass
        
        check_stop()
        log(f"🎯 Clicking apply button (applyBtn) in modal for lottery #{lottery_number}...", 'info')
        apply_btn = wait.until(EC.element_to_be_clickable((By.ID, "applyBtn")))
        _human_like_scroll_to_element(driver, apply_btn)
        time.sleep(random.uniform(0.3, 0.5))
        _human_like_click(driver, apply_btn)
        log(f"✅ Application submitted successfully for lottery #{lottery_number}!", 'success')
        
        # Wait for confirmation and page response
        for _ in range(5):
            check_stop()
            time.sleep(1)
        
        # Immediately check for pop04/pop05 errors after apply button click
        check_stop()
        log("🔍 Checking for pop04/pop05 errors immediately after apply button click...", 'info')
        pop_reload_needed_after_apply = _check_and_handle_pop_exceptions(driver, wait, max_reload_attempts=5)
        if pop_reload_needed_after_apply:
            log("⚠️ Pop04/pop05 error detected after apply button click. Page reloaded - will restart from first lottery.", 'warning')
            # Return special value to indicate reload is needed
            return 'reload_needed'
        
        # Check if page was reloaded or navigated after submission
        check_stop()
        current_url_after_submit = driver.current_url
        log(f"📍 Current URL after submission: {current_url_after_submit}", 'info')
        
        # If we're not on apply page, navigate back to it
        if "apply.html" not in current_url_after_submit:
            log(f"⚠️ Not on apply page after submission. Navigating back to apply page...", 'warning')
            try:
                driver.get(APPLY_URL)
                log(f"✅ Navigated back to apply page", 'success')
                # Wait for page to load
                for _ in range(5):
                    check_stop()
                    time.sleep(1)
            except Exception as e:
                log(f"⚠️ Could not navigate back to apply page: {e}. Continuing anyway...", 'warning')
        else:
            # Check if modal is still open and close it if needed
            check_stop()
            try:
                pop01_modal = driver.find_elements(By.ID, "pop01")
                if pop01_modal and pop01_modal[0].is_displayed():
                    log(f"🔔 Modal (pop01) is still open after submission. Closing modal...", 'info')
                    try:
                        # Try to find and click close button or overlay
                        close_buttons = driver.find_elements(By.XPATH, '//*[@id="pop01"]//button[contains(@class, "close")] | //*[@id="pop01"]//a[contains(@class, "close")] | //*[@id="pop01"]//*[contains(@onclick, "close")]')
                        if close_buttons:
                            _human_like_click(driver, close_buttons[0])
                            log(f"✅ Modal closed via close button", 'success')
                        else:
                            # Try clicking outside modal or ESC key
                            driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                            log(f"✅ Modal closed via ESC key", 'success')
                        time.sleep(1)
                    except Exception as e:
                        log(f"⚠️ Could not close modal: {e}. Checking for pop04 exception message...", 'warning')
                        # Check for pop04 exception message instead of automatically reloading
                        check_stop()
                        pop04_reload_needed = _check_and_handle_pop_exceptions(driver, wait, max_reload_attempts=5)
                        if not pop04_reload_needed:
                            # If no pop04 exception, try to navigate back to apply page only if not already there
                            try:
                                if "apply.html" not in driver.current_url:
                                    driver.get(APPLY_URL)
                                    log(f"✅ Navigated back to apply page to clear modal", 'success')
                                    for _ in range(3):
                                        check_stop()
                                        time.sleep(1)
                            except Exception as e2:
                                log(f"⚠️ Could not navigate to apply page: {e2}. Continuing anyway...", 'warning')
            except Exception as e:
                # Modal might have auto-closed or doesn't exist
                log(f"ℹ️ Modal check completed (might be auto-closed): {e}", 'info')
        
        log(f"🎉 Lottery #{lottery_number} entry process completed successfully!", 'success')
        
        # Check for pop04 exception message "意図しない例外が発生しました。" after lottery processing
        check_stop()
        log("🔍 Checking for pop04 exception message after lottery processing...", 'info')
        pop04_reload_needed = _check_and_handle_pop_exceptions(driver, wait, max_reload_attempts=5)
        if pop04_reload_needed:
            log("✅ Pop04 exception handled and page reloaded", 'success')
        
        return True  # Return True on success
        
    except StopIteration:
        log(f"⏹️ Lottery #{lottery_number} entry process stopped by user", 'warning')
        raise
    except Exception as e:
        log(f"❌ Error processing lottery #{lottery_number}: {e}", 'error')
        import traceback
        traceback.print_exc()
        return False  # Return False on failure (don't raise to allow continuing with next lottery)

# Load data from Excel file row by row
def load_data_from_excel():
    chrome_options = Options()
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option('useAutomationExtension', False)
    # Suppress Chrome warnings and errors
    chrome_options.add_argument('--log-level=3')  # Only show fatal errors
    chrome_options.add_argument('--disable-logging')
    chrome_options.add_argument('--disable-gpu-logging')
    chrome_options.add_argument('--disable-background-networking')  # Disable GCM/background services
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
    })
    wait = WebDriverWait(driver, 30)
    
    try:
        excel_file = "V0vMwOh.xlsx"
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=1, values_only=False):
            user_email = row[0].value  # Column A is the first column (index 0)
            user_password = row[1].value if len(row) > 1 else None  # Column B is the second column (index 1)
            
            # Email is required from Excel file (not from .env)
            if not user_email:
                log(f"⚠️ Skipping row with empty email", 'warning')
                continue
            
            global EMAIL, PASSWORD
            EMAIL = user_email  # Update module-level EMAIL variable (required from Excel)
            
            # Update PASSWORD from Excel if available, otherwise use env variable or raise error
            if user_password:
                PASSWORD = user_password
            elif PASSWORD is None:
                raise ValueError("PASSWORD is not set. Please set it in .env file or include it in column B of the Excel file.")
            
            log(f"📧 Processing email: {user_email}", 'info')
            lottery_begin(driver, wait)

        workbook.close()
    except Exception as e:
        log(f"❌ Error: {e}", 'error')
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()
        input("\nPress Enter to stop the project...")




if __name__ == "__main__":
    load_data_from_excel()
